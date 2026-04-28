import openpyxl

def fill_indicators(ws: openpyxl.worksheet, config: list, data_sheet: str = 'DATA') -> None:
    """
    Fill the indicators in the Excel file with formulas.

    Args:
        path_file (str): Path to the Excel file.
        data_sheet (str, optional): Name of the data sheet. Defaults to 'DATA'.
    """

    for item in config:
        cell = f'E{item["row"]}'
        
        if item['formule'] == 'COUNTIF':
            col, val = item['args'][0]
            ws[cell] = f'=COUNTIF({data_sheet}!{col}:{col}, "{val}")'
            
        elif item['formule'] == 'COUNTIFS':
            criteria = [f'{data_sheet}!{col}:{col}, "{val}"' for col, val in item['args']]
            ws[cell] = f'=COUNTIFS({", ".join(criteria)})'
            
        elif item['formule'] == 'SUM':
            ws[cell] = f'=SUM({item["args"][0]})'

    #wb.save(path_file) à mettre dans main
    #wb.close() à mettre dans main